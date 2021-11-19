import pprint
import json
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
import time

start_time = time.time()

wb = load_workbook('../price.xlsx', data_only=True)

sheet = wb['specification']

artic = sheet['C']
color = sheet['D']
name = sheet['E']
manufacturer = sheet['F']
country = sheet['G']
mark = sheet['H']
material_up = sheet['I']
material_down = sheet['J']
sole = sheet['K']
type_material = sheet['L']
gender = sheet['M']
size = sheet['N']
lay = sheet['O']
tvned = sheet['P']
unit = sheet['Q']
multiplier = sheet['R']
doc = sheet['AZ']
date_start = sheet['BB']
date_close = sheet['BC']
filial = sheet['BD']
status_dov = sheet['BE']
heel_height = sheet['BO']

x = 1
dcty = {}

for i in range(2, sheet.max_row):
    dct = {}
    dct[artic[0].value] = (sheet.cell(row=i, column=column_index_from_string('C')).value)
    dct[color[0].value] = (sheet.cell(row=i, column=column_index_from_string('D')).value)
    dct[name[0].value] = (sheet.cell(row=i, column=column_index_from_string('E')).value)
    dct[manufacturer[0].value] = (sheet.cell(row=i, column=column_index_from_string('F')).value)
    dct[country[0].value] = (sheet.cell(row=i, column=column_index_from_string('G')).value)
    dct[mark[0].value] = (sheet.cell(row=i, column=column_index_from_string('H')).value)
    dct[material_up[0].value] = (sheet.cell(row=i, column=column_index_from_string('I')).value)
    dct[material_down[0].value] = (sheet.cell(row=i, column=column_index_from_string('J')).value)
    dct[sole[0].value] = (sheet.cell(row=i, column=column_index_from_string('K')).value)
    dct[type_material[0].value] = (sheet.cell(row=i, column=column_index_from_string('L')).value)
    dct[gender[0].value] = (sheet.cell(row=i, column=column_index_from_string('M')).value)
    dct[size[0].value] = (sheet.cell(row=i, column=column_index_from_string('N')).value)
    dct[lay[0].value] = (sheet.cell(row=i, column=column_index_from_string('O')).value)
    dct[tvned[0].value] = (sheet.cell(row=i, column=column_index_from_string('P')).value)
    dct[unit[0].value] = (sheet.cell(row=i, column=column_index_from_string('Q')).value)
    dct[multiplier[0].value] = (sheet.cell(row=i, column=column_index_from_string('R')).value)
    dct[doc[0].value] = (sheet.cell(row=i, column=column_index_from_string('AZ')).value)
    # dct[date_start[0].value] = (sheet.cell(row=i, column=column_index_from_string('BB')).value)
    # dct[date_close[0].value] = (sheet.cell(row=i, column=column_index_from_string('BC')).value)
    dct[filial[0].value] = (sheet.cell(row=i, column=column_index_from_string('BD')).value)
    dct[status_dov[0].value] = (sheet.cell(row=i, column=column_index_from_string('BE')).value)
    dct[heel_height[0].value] = (sheet.cell(row=i, column=column_index_from_string('BO')).value)

    if not str(sheet.cell(row=i, column=column_index_from_string('P')).value)[0:4] in dcty:
        dcty[str(sheet.cell(row=i, column=column_index_from_string('P')).value)[0:4]] = []
        dcty[str(sheet.cell(row=i, column=column_index_from_string('P')).value)[0:4]].append(dct)
    else:
        dcty[str(sheet.cell(row=i, column=column_index_from_string('P')).value)[0:4]].append(dct)

    with open('../json_file_002.json', 'w', encoding='utf-8') as file:
        json.dump(dcty, file, ensure_ascii=False, indent=4)

# pprint.pprint(dcty)
print('---%s seconds ---' % round((time.time() - start_time), 3))

